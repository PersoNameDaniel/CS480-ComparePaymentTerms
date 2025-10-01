"""Excel processing module for reading XLSX files and QuickBooks integration.

This module provides functions to read Excel files, specifically payment terms,
and integrate with QuickBooks Desktop via COM API.
"""

from dataclasses import dataclass
from typing import Any

import win32com.client
from openpyxl import load_workbook


@dataclass
class PaymentTerm:
    """Represents a payment term with name and ID.

    The ID is stored in QuickBooks' StdDiscountDays field for matching purposes.
    """

    name: str
    term_id: int  # Excel ID, stored in QB's StdDiscountDays field


@dataclass
class TermComparison:
    """Results of comparing Excel and QuickBooks payment terms."""

    same_id_diff_name: list[tuple[str, str, int]]  # (excel_name, qb_name, term_id)
    only_in_excel: list[PaymentTerm]  # Terms to add to QB
    only_in_qb: list[PaymentTerm]  # Terms in QB but not Excel
    matching_count: int  # Count of terms with same ID and name


def read_payment_terms(file_path: str) -> list[PaymentTerm]:
    """Read payment terms from the specified Excel file.

    Expected Excel format:
    - Must contain a sheet named 'payment_terms'
    - Column A: Payment term names (strings)
    - Column B: Term ID (integers)
    - Row 1 should contain headers (will be skipped)
    - Data starts from row 2

    Args:
        file_path (str): Path to the Excel file containing payment terms (.xlsx format)

    Returns:
        list[PaymentTerm]: List of payment terms with name and term_id.
                          Empty list if no valid payment terms found.

    Raises:
        No exceptions need to be manually raised - let openpyxl handle file/sheet errors
    """
    workbook = load_workbook(file_path, read_only=True)
    sheet = workbook["payment_terms"]
    payment_terms = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        name = row[0]
        term_id = row[1]

        # Skip rows with missing data
        if name is None or term_id is None:
            continue

        # Convert and validate data
        try:
            name_str = str(name).strip()
            term_id_int = int(term_id)
            if name_str:  # Only add if name is not empty
                payment_terms.append(PaymentTerm(name=name_str, term_id=term_id_int))
        except (ValueError, TypeError):
            # Skip rows with invalid data
            continue

    return payment_terms


def connect_to_quickbooks() -> Any:
    """Connect to QuickBooks Desktop via COM API.

    This function establishes a connection to QuickBooks Desktop using the
    QBXML Request Processor COM interface. QuickBooks Desktop must be running
    with a company file open.

    Returns:
        tuple[Any, Any]: A tuple containing (qb_app, session)
            - qb_app: COM object for QuickBooks application interface
            - session: Session ticket for the current QB connection

    Raises:
        No exceptions need to be manually raised - let win32com handle COM errors

    Security Notes:
        - User may need to grant permission in QuickBooks for first-time access
        - QuickBooks may prompt user to allow external application access
    """
    try:
        qb_app = win32com.client.Dispatch("QBXMLRP2.RequestProcessor")
        qb_app.OpenConnection("", "Payment Terms Import")
        session = qb_app.BeginSession("", 2)  # 2 = qbFileOpenDoNotCare
        return qb_app, session
    except Exception as e:
        print(f"QuickBooks connection error: {str(e)}")
        raise


def get_qb_payment_terms() -> list[PaymentTerm]:
    """Read existing payment terms from QuickBooks Desktop.

    Queries QuickBooks for all StandardTerms and returns them as PaymentTerm objects.
    The term_id is read from the StdDiscountDays field.

    Returns:
        list[PaymentTerm]: List of payment terms from QuickBooks with name and term_id
                          (where term_id comes from StdDiscountDays field).

    Raises:
        RuntimeError: If connection to QuickBooks fails.
    """
    qb_app = None
    session = None
    try:
        qb_app, session = connect_to_quickbooks()

        # Create QBXML query to get all standard terms
        qbxml_query = """<?xml version="1.0" encoding="utf-8"?>
<?qbxml version="13.0"?>
<QBXML>
    <QBXMLMsgsRq onError="continueOnError">
        <StandardTermsQueryRq>
        </StandardTermsQueryRq>
    </QBXMLMsgsRq>
</QBXML>"""

        response = qb_app.ProcessRequest(session, qbxml_query)

        # Parse the response
        payment_terms = []
        if "<StandardTermsRet>" in response:
            # Extract each term from the response
            import xml.etree.ElementTree as ET

            root = ET.fromstring(response)
            for term_ret in root.findall(".//StandardTermsRet"):
                name_elem = term_ret.find("Name")
                discount_days_elem = term_ret.find("StdDiscountDays")

                if name_elem is not None and discount_days_elem is not None:
                    name = name_elem.text
                    if name is not None and discount_days_elem.text is not None:
                        try:
                            term_id = int(discount_days_elem.text)
                            payment_terms.append(PaymentTerm(name=name, term_id=term_id))
                        except (ValueError, TypeError):
                            # Skip terms without valid discount days
                            continue

        return payment_terms

    except Exception as e:
        raise RuntimeError(f"Failed to read QuickBooks payment terms: {str(e)}") from e
    finally:
        # Clean up connection
        if qb_app is not None and session is not None:
            qb_app.EndSession(session)
        if qb_app is not None:
            qb_app.CloseConnection()


def compare_payment_terms(
    excel_terms: list[PaymentTerm], qb_terms: list[PaymentTerm]
) -> TermComparison:
    """Compare Excel and QuickBooks payment terms by ID.

    Args:
        excel_terms (list[PaymentTerm]): Payment terms from Excel file.
            Example: [PaymentTerm(name="Net 30", term_id=30),
                     PaymentTerm(name="Net 60", term_id=60)]
        qb_terms (list[PaymentTerm]): Payment terms from QuickBooks.
            Example: [PaymentTerm(name="Net 30", term_id=30),
                     PaymentTerm(name="2% 10 Net 30", term_id=10)]

    Returns:
        TermComparison: A dataclass containing comparison results:
            - same_id_diff_name (list[tuple[str, str, int]]): Terms with matching IDs but
              different names. Each tuple contains (excel_name, qb_name, term_id).
              Example: [("Net 15", "Net 15 Days", 15)]
            - only_in_excel (list[PaymentTerm]): Terms present in Excel but not in QB.
              These need to be added to QuickBooks.
              Example: [PaymentTerm(name="Net 60", term_id=60)]
            - only_in_qb (list[PaymentTerm]): Terms present in QB but not in Excel.
              Example: [PaymentTerm(name="Due on Receipt", term_id=0)]
            - matching_count (int): Count of terms with identical ID and name.
              Example: 5
    """
    qb_terms_dict = {term.term_id: term.name for term in qb_terms}
    excel_terms_dict = {term.term_id: term.name for term in excel_terms}

    same_id_diff_name = []
    only_in_excel = []
    only_in_qb = []
    matching_count = 0

    # Check Excel terms against QB terms
    for excel_term in excel_terms:
        qb_name = qb_terms_dict.get(excel_term.term_id)
        if qb_name is None:
            # Term ID not found in QB - needs to be added
            only_in_excel.append(excel_term)
        elif qb_name != excel_term.name:
            # Term ID found but names differ
            same_id_diff_name.append((excel_term.name, qb_name, excel_term.term_id))
        else:
            # Exact match
            matching_count += 1

    # Check QB terms against Excel terms to find those only in QB
    for qb_term in qb_terms:
        if qb_term.term_id not in excel_terms_dict:
            only_in_qb.append(qb_term)

    return TermComparison(
        same_id_diff_name=same_id_diff_name,
        only_in_excel=only_in_excel,
        only_in_qb=only_in_qb,
        matching_count=matching_count,
    )


def create_payment_terms_batch_qbxml(payment_terms: list[PaymentTerm]) -> str:
    """Create QBXML for adding multiple payment terms in a batch.

    This function generates a well-formed QBXML document containing multiple
    StandardTermsAddRq requests that can be sent to QuickBooks Desktop in a
    single batch operation. The term_id is stored in StdDiscountDays field.

    Args:
        payment_terms (list[PaymentTerm]): List of payment terms to create.
                                         Each PaymentTerm must have name and term_id.

    Returns:
        str: Complete QBXML string ready to send to QuickBooks Desktop.
             Contains XML declaration, QBXML root, and multiple StandardTermsAddRq elements.

    Raises:
        AttributeError: If PaymentTerm objects are missing required attributes
        TypeError: If payment_terms is not a list or contains invalid objects
    """
    term_requests = []
    for term in payment_terms:
        # Escape XML special characters
        name = term.name.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
        term_request = f"""        <StandardTermsAddRq>
            <StandardTermsAdd>
                <Name>{name}</Name>
                <StdDueDays >30</StdDueDays >
                <StdDiscountDays >{term.term_id}</StdDiscountDays >
            </StandardTermsAdd>
        </StandardTermsAddRq>"""
        term_requests.append(term_request)

    qbxml = f"""<?xml version="1.0" encoding="utf-8"?>
<?qbxml version="13.0"?>
<QBXML>
    <QBXMLMsgsRq onError="continueOnError">
{chr(10).join(term_requests)}
    </QBXMLMsgsRq>
</QBXML>"""

    return qbxml


def save_payment_terms_to_quickbooks(payment_terms: list[PaymentTerm]) -> list[str]:
    """Save payment terms to QuickBooks Desktop.

    This function connects to QuickBooks, sends a batch QBXML request to create
    multiple payment terms, parses the response, and returns the list of
    successfully created terms.

    Args:
        payment_terms (list[PaymentTerm]): List of payment terms to save to QuickBooks.
                                         Each term must have valid name and term_id.

    Returns:
        list[str]: List of payment term names that were successfully created.
                  May be shorter than input list if some terms failed or already exist.

    Raises:
        RuntimeError: If connection to QuickBooks fails
    """
    if not payment_terms:
        return []

    try:
        qb_app, session = connect_to_quickbooks()

        qbxml = create_payment_terms_batch_qbxml(payment_terms)
        response = qb_app.ProcessRequest(session, qbxml)

        # Parse response
        import xml.etree.ElementTree as ET

        created_terms = []
        root = ET.fromstring(response)

        for add_rs in root.findall(".//StandardTermsAddRs"):
            status_code = add_rs.get("statusCode")
            if status_code == "0":
                # Successfully created
                name_elem = add_rs.find(".//Name")
                if name_elem is not None and name_elem.text is not None:
                    created_terms.append(name_elem.text)
            elif status_code == "3100":
                # Term already exists - skip silently
                pass
            else:
                # Other error
                status_msg = add_rs.get("statusMessage", "Unknown error")
                print(f"Warning: Failed to create term: {status_msg}")

        # Cleanup
        qb_app.EndSession(session)
        qb_app.CloseConnection()

        return created_terms

    except Exception as e:
        raise RuntimeError(f"Failed to connect to QuickBooks: {str(e)}") from e


def process_payment_terms(file_path: str) -> TermComparison:
    """Read payment terms from Excel, compare with QuickBooks, and synchronize.

    This is the main orchestration function that:
    1. Reads payment terms from Excel
    2. Reads existing payment terms from QuickBooks
    3. Compares them by ID (stored in QB's StdDiscountDays field)
    4. Prints differences and matches
    5. Adds new terms from Excel to QuickBooks

    Args:
        file_path (str): Path to the Excel file containing payment terms (.xlsx format).
                        File must contain a 'payment_terms' sheet with Name and ID columns.

    Returns:
        TermComparison: Comparison results containing differences and matches

    Raises:
        ValueError: If no payment terms found in the Excel file
    """
    # Read Excel payment terms
    excel_terms = read_payment_terms(file_path)
    if not excel_terms:
        raise ValueError("No payment terms found in Excel file")

    print(f"Found {len(excel_terms)} payment terms in Excel")

    # Read QuickBooks payment terms
    print("Reading payment terms from QuickBooks...")
    qb_terms = get_qb_payment_terms()
    print(f"Found {len(qb_terms)} payment terms in QuickBooks")

    # Compare terms
    comparison = compare_payment_terms(excel_terms, qb_terms)

    # Print results
    print("\n=== Comparison Results ===")

    # Print matching terms count
    print(f"\nMatching terms (same ID and name): {comparison.matching_count}")

    # Print terms with same ID but different names
    if comparison.same_id_diff_name:
        print(f"\nTerms with same ID but different names ({len(comparison.same_id_diff_name)}):")
        for excel_name, qb_name, term_id in comparison.same_id_diff_name:
            print(f"  ID {term_id}: Excel='{excel_name}' vs QB='{qb_name}'")
    else:
        print("\nNo terms with same ID but different names")

    # Print terms only in QB
    if comparison.only_in_qb:
        print(f"\nTerms in QuickBooks but not in Excel ({len(comparison.only_in_qb)}):")
        for term in comparison.only_in_qb:
            print(f"  ID {term.term_id}: {term.name}")
    else:
        print("\nNo terms found only in QuickBooks")

    # Add terms from Excel to QB
    if comparison.only_in_excel:
        print(f"\nAdding {len(comparison.only_in_excel)} new terms to QuickBooks...")
        for term in comparison.only_in_excel:
            print(f"  - {term.name} (ID: {term.term_id})")

        created_terms = save_payment_terms_to_quickbooks(comparison.only_in_excel)
        print(f"\nSuccessfully created {len(created_terms)} terms in QuickBooks")
    else:
        print("\nNo new terms to add to QuickBooks")

    return comparison
